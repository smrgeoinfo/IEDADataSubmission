"""
Lightweight file inspectors for bundle introspection.

Each inspector extracts metadata from a specific file type and returns
a dict with a common structure suitable for pre-populating JSON-LD
metadata fields (variableMeasured, file descriptions, etc.).

All third-party dependencies are optional; inspectors gracefully degrade
when the required library is not installed.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Optional imports
# ---------------------------------------------------------------------------

try:
    import openpyxl
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import h5py
    import numpy as np

    _HAS_H5PY = True
except ImportError:
    _HAS_H5PY = False

try:
    import netCDF4

    _HAS_NETCDF4 = True
except ImportError:
    _HAS_NETCDF4 = False

try:
    from pypdf import PdfReader

    _HAS_PYPDF = True
except ImportError:
    _HAS_PYPDF = False


# ---------------------------------------------------------------------------
# Common helpers
# ---------------------------------------------------------------------------

def _safe_float(val: Any) -> Optional[float]:
    """Convert a value to float, returning None on failure."""
    if val is None:
        return None
    try:
        f = float(val)
        if f != f:  # NaN check
            return None
        return f
    except (ValueError, TypeError, OverflowError):
        return None


def _detect_delimiter(lines: List[str]) -> str:
    """Detect the most likely CSV delimiter from sample lines."""
    delimiters = [",", "\t", ";", "|"]
    sample = [l for l in lines[:20] if l.strip() and not l.strip().startswith("#")][:5]
    if not sample:
        return ","
    scores: Dict[str, float] = {}
    for d in delimiters:
        counts = [line.count(d) for line in sample]
        avg = sum(counts) / len(counts)
        if avg > 0:
            var = sum((c - avg) ** 2 for c in counts) / len(counts)
            scores[d] = avg / (1 + var)
    return max(scores, key=scores.get) if scores else ","


# ---------------------------------------------------------------------------
# CSV / delimited text inspector
# ---------------------------------------------------------------------------

def inspect_csv(file_path: str, max_rows: int = 200) -> Dict[str, Any]:
    """
    Inspect a CSV/TSV file and return column metadata.

    Returns
    -------
    dict with keys: columns, row_count, delimiter, size, mime_type, warnings
    """
    result: Dict[str, Any] = {
        "columns": [],
        "row_count": 0,
        "delimiter": ",",
        "size": os.path.getsize(file_path),
        "mime_type": "text/csv",
        "warnings": [],
    }

    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            raw = f.read()
    except Exception as e:
        result["warnings"].append(f"Could not read file: {e}")
        return result

    lines = raw.splitlines()
    if not lines:
        result["warnings"].append("Empty file")
        return result

    delimiter = _detect_delimiter(lines)
    result["delimiter"] = delimiter

    # Skip comment lines
    data_lines = [l for l in lines if l.strip() and not l.strip().startswith("#")]
    if not data_lines:
        return result

    reader = csv.reader(data_lines, delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return result

    header = rows[0]
    data_rows = rows[1: max_rows + 1]
    result["row_count"] = len(rows) - 1  # total data rows (approximate for large files)

    for col_idx, label in enumerate(header):
        label = label.strip()
        if not label:
            continue

        col_info: Dict[str, Any] = {
            "index": col_idx,
            "label": label,
            "name": label.replace("_", " ").strip(),
            "data_type": "string",
            "min_val": None,
            "max_val": None,
            "sample_values": [],
        }

        # Collect values for this column
        values = []
        for row in data_rows:
            if col_idx < len(row) and row[col_idx].strip():
                values.append(row[col_idx].strip())

        col_info["sample_values"] = values[:5]

        # Detect numeric type
        numeric_vals: List[float] = []
        for v in values:
            f = _safe_float(v.replace(",", ""))
            if f is not None:
                numeric_vals.append(f)

        if numeric_vals and len(numeric_vals) > len(values) * 0.5:
            all_int = all(v == int(v) for v in numeric_vals)
            col_info["data_type"] = "integer" if all_int else "decimal"
            col_info["min_val"] = min(numeric_vals)
            col_info["max_val"] = max(numeric_vals)

        result["columns"].append(col_info)

    return result


# ---------------------------------------------------------------------------
# Excel inspector
# ---------------------------------------------------------------------------

def inspect_excel(file_path: str, max_rows: int = 200) -> Dict[str, Any]:
    """
    Inspect an Excel workbook and return column metadata.

    Returns
    -------
    dict with keys: columns, row_count, sheet_count, sheet_names,
                    size, mime_type, warnings
    """
    result: Dict[str, Any] = {
        "columns": [],
        "row_count": 0,
        "sheet_count": 0,
        "sheet_names": [],
        "size": os.path.getsize(file_path),
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "warnings": [],
    }

    if not _HAS_OPENPYXL:
        result["warnings"].append("openpyxl not installed; cannot inspect Excel files")
        return result

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        result["warnings"].append(f"Could not open workbook: {e}")
        return result

    try:
        result["sheet_count"] = len(wb.sheetnames)
        result["sheet_names"] = list(wb.sheetnames)

        # Inspect first sheet
        sheet = wb[wb.sheetnames[0]]
        rows: List[list] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) > max_rows + 10:
                break

        if not rows:
            wb.close()
            return result

        # Detect header row
        header_idx = _detect_excel_header(rows)
        header = rows[header_idx] if header_idx < len(rows) else rows[0]
        data_rows = rows[header_idx + 1:]
        result["row_count"] = len(data_rows)

        for col_idx, hval in enumerate(header):
            if hval is None:
                continue
            label = str(hval).strip()
            if not label:
                continue

            col_info: Dict[str, Any] = {
                "index": col_idx,
                "label": label,
                "name": label.replace("_", " ").strip(),
                "data_type": "string",
                "min_val": None,
                "max_val": None,
                "sample_values": [],
            }

            # Collect values
            numeric_vals: List[float] = []
            samples: List[str] = []
            for row in data_rows[:max_rows]:
                if col_idx < len(row) and row[col_idx] is not None:
                    val = row[col_idx]
                    samples.append(str(val)[:100])
                    f = _safe_float(val)
                    if f is not None:
                        numeric_vals.append(f)

            col_info["sample_values"] = samples[:5]

            if numeric_vals and len(numeric_vals) > len(samples) * 0.5:
                all_int = all(v == int(v) for v in numeric_vals)
                col_info["data_type"] = "integer" if all_int else "decimal"
                col_info["min_val"] = min(numeric_vals)
                col_info["max_val"] = max(numeric_vals)

            result["columns"].append(col_info)

        wb.close()
    except Exception as e:
        result["warnings"].append(f"Error inspecting workbook: {e}")
        try:
            wb.close()
        except Exception:
            pass

    return result


def _detect_excel_header(rows: List[list]) -> int:
    """Return the index of the probable header row."""
    for i, row in enumerate(rows[:5]):
        non_empty = [c for c in row if c is not None]
        if len(non_empty) >= 2:
            text_count = sum(1 for c in non_empty if isinstance(c, str))
            if text_count >= len(non_empty) * 0.5:
                return i
    return 0


# ---------------------------------------------------------------------------
# HDF5 inspector
# ---------------------------------------------------------------------------

def inspect_hdf5(file_path: str) -> Dict[str, Any]:
    """
    Inspect an HDF5 file and return variable metadata.

    Returns
    -------
    dict with keys: variables, size, mime_type, warnings
    """
    result: Dict[str, Any] = {
        "variables": [],
        "size": os.path.getsize(file_path),
        "mime_type": "application/x-hdf5",
        "warnings": [],
    }

    if not _HAS_H5PY:
        result["warnings"].append("h5py not installed; cannot inspect HDF5 files")
        return result

    try:
        with h5py.File(file_path, "r") as f:
            _collect_hdf5_vars(f, result["variables"])
    except Exception as e:
        result["warnings"].append(f"Error reading HDF5 file: {e}")

    return result


def _collect_hdf5_vars(group, variables: list, prefix: str = ""):
    """Recursively collect datasets from an HDF5 group."""
    for key in group:
        path = f"{prefix}/{key}" if prefix else key
        item = group[key]
        if isinstance(item, h5py.Dataset):
            var_info: Dict[str, Any] = {
                "name": key,
                "path": path,
                "description": _get_hdf5_attr(item, "long_name", ""),
                "unit": _get_hdf5_attr(item, "units", ""),
                "data_type": str(item.dtype),
                "shape": list(item.shape),
                "min_val": None,
                "max_val": None,
            }
            # Try to compute min/max for numeric data (small datasets only)
            if item.dtype.kind in ("f", "i", "u") and item.size < 10_000_000:
                try:
                    data = item[()]
                    valid = data[np.isfinite(data)] if np.issubdtype(data.dtype, np.floating) else data
                    if valid.size > 0:
                        var_info["min_val"] = float(np.min(valid))
                        var_info["max_val"] = float(np.max(valid))
                except Exception:
                    pass
            variables.append(var_info)
        elif isinstance(item, h5py.Group):
            _collect_hdf5_vars(item, variables, prefix=path)


def _get_hdf5_attr(dataset, attr_name: str, default: str = "") -> str:
    """Get an attribute from an HDF5 dataset, returning default if missing."""
    try:
        val = dataset.attrs.get(attr_name)
        if val is not None:
            if isinstance(val, bytes):
                return val.decode("utf-8", errors="replace")
            return str(val)
    except Exception:
        pass
    return default


# ---------------------------------------------------------------------------
# NetCDF inspector
# ---------------------------------------------------------------------------

def inspect_netcdf(file_path: str) -> Dict[str, Any]:
    """
    Inspect a NetCDF file and return variable metadata.

    Returns
    -------
    dict with keys: variables, dimensions, size, mime_type, warnings
    """
    result: Dict[str, Any] = {
        "variables": [],
        "dimensions": [],
        "size": os.path.getsize(file_path),
        "mime_type": "application/x-netcdf",
        "warnings": [],
    }

    if not _HAS_NETCDF4:
        result["warnings"].append("netCDF4 not installed; cannot inspect NetCDF files")
        return result

    try:
        ds = netCDF4.Dataset(file_path, "r")
    except Exception as e:
        result["warnings"].append(f"Error reading NetCDF file: {e}")
        return result

    try:
        # Dimensions
        for dim_name, dim in ds.dimensions.items():
            result["dimensions"].append({
                "name": dim_name,
                "size": len(dim),
                "unlimited": dim.isunlimited(),
            })

        # Variables
        for var_name, var in ds.variables.items():
            var_info: Dict[str, Any] = {
                "name": var_name,
                "description": getattr(var, "long_name", ""),
                "unit": getattr(var, "units", ""),
                "data_type": str(var.dtype),
                "dimensions": list(var.dimensions),
                "shape": list(var.shape),
                "min_val": None,
                "max_val": None,
            }
            # Min/max for manageable-size numeric vars
            if var.dtype.kind in ("f", "i", "u") and var.size < 10_000_000:
                try:
                    data = var[:]
                    if hasattr(data, "compressed"):
                        data = data.compressed()  # handle masked arrays
                    valid = data[np.isfinite(data)] if np.issubdtype(data.dtype, np.floating) else data
                    if valid.size > 0:
                        var_info["min_val"] = float(np.min(valid))
                        var_info["max_val"] = float(np.max(valid))
                except Exception:
                    pass
            result["variables"].append(var_info)

        ds.close()
    except Exception as e:
        result["warnings"].append(f"Error inspecting NetCDF file: {e}")
        try:
            ds.close()
        except Exception:
            pass

    return result


# ---------------------------------------------------------------------------
# PDF inspector — extract opening text for description
# ---------------------------------------------------------------------------

_MAX_PDF_CHARS = 2000  # max chars to extract


def inspect_pdf(file_path: str) -> Dict[str, Any]:
    """
    Inspect a PDF file and extract opening text for use as a description.

    Returns
    -------
    dict with keys: description, page_count, size, mime_type, warnings
    """
    result: Dict[str, Any] = {
        "description": "",
        "page_count": 0,
        "size": os.path.getsize(file_path),
        "mime_type": "application/pdf",
        "warnings": [],
    }

    if not _HAS_PYPDF:
        result["warnings"].append("pypdf not installed; cannot inspect PDF files")
        return result

    try:
        reader = PdfReader(file_path)
        result["page_count"] = len(reader.pages)

        # Extract text from first few pages
        text_parts: List[str] = []
        chars = 0
        for page in reader.pages[:3]:
            try:
                t = page.extract_text() or ""
                text_parts.append(t)
                chars += len(t)
                if chars >= _MAX_PDF_CHARS:
                    break
            except Exception:
                continue

        full_text = " ".join(text_parts)
        # Collapse whitespace/line-feeds for clean UI word wrapping
        full_text = re.sub(r'\s+', ' ', full_text).strip()
        # Truncate to reasonable length
        if len(full_text) > _MAX_PDF_CHARS:
            full_text = full_text[:_MAX_PDF_CHARS].rsplit(" ", 1)[0] + "..."

        result["description"] = full_text

    except Exception as e:
        result["warnings"].append(f"Error reading PDF: {e}")

    return result


# ---------------------------------------------------------------------------
# Plain text inspector — extract opening lines for description
# ---------------------------------------------------------------------------

_MAX_TEXT_CHARS = 2000


def inspect_text(file_path: str) -> Dict[str, Any]:
    """
    Inspect a plain text file and extract opening content for description.

    Returns
    -------
    dict with keys: description, line_count, size, mime_type, warnings
    """
    result: Dict[str, Any] = {
        "description": "",
        "line_count": 0,
        "size": os.path.getsize(file_path),
        "mime_type": "text/plain",
        "warnings": [],
    }

    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read(_MAX_TEXT_CHARS + 500)
        result["line_count"] = content.count("\n") + 1
        desc = content[:_MAX_TEXT_CHARS]
        # Collapse whitespace/line-feeds for clean UI word wrapping
        desc = re.sub(r'\s+', ' ', desc).strip()
        if len(content) > _MAX_TEXT_CHARS:
            desc = desc.rsplit(" ", 1)[0] + "..."
        result["description"] = desc
    except Exception as e:
        result["warnings"].append(f"Error reading text file: {e}")

    return result
